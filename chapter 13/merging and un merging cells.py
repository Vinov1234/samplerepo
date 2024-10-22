from openpyxl import Workbook
wb = Workbook()
sheet = wb.active
sheet.merge_cells("G1:M1")
wb.save(filename = "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\merge.xlsx")
sheet.unmerge_cells("G1:M1")
wb.save(filename =  "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\merge.xlsx")

