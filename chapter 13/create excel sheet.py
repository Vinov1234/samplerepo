from openpyxl import Workbook
wb = Workbook()
wb.active
sheet = wb.active
wb.save(filename = "C:\\Users\pc\\Documents\\Automated the boring stuff\\chapter 13\\par2.xlsx")
sheet1 = wb.create_sheet()
sheet2 = wb.create_sheet("Wbs1")
wb.save(filename = "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\par2_1.xlsx")
sheet3 = wb.create_sheet("WbS2", 0)
sheet4 = wb.create_sheet("Wbs2", 1)
sheet5 = wb.create_sheet("Wbs3", 2)
wb.save(filename = "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\par2_1_1.xlsx")
sheet1.title = "new title"
wb.save(filename = "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\par2_1_1.xlsx")
sheet3.sheet_properties.tabColor = "1072BA"
wb.save(filename = "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\par2_1_1.xlsx")
