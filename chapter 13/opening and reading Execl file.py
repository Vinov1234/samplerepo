import openpyxl
path =  "C:\\Users\pc\\Documents\\Automated the boring stuff\\chapter 13\\part5.xlsx"
wb_file = openpyxl.Load_Workbook(path)
sheet = wb_file.active
sheet.cell(row = 3, column = 3)
val1 = sheet.cell(row = 3,column = 3)
val1.value
sheet.max_row
sheet. max_column
max_col = sheet.max_column
max_col
for i in range (1,max_col + 1):
    cell_val = sheet.cell(row = 1,column = i)
    print(cell_val.value)

for i in range (1,max_rows+1):
    cell_val = sheet.cell(row = i, column = 1)
    print(cell_val.value)
