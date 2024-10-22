import openpyxl
wb = openpyxl.load_workbook(filename = "C:\\Users\\pc\\Documents\\Automated the boring stuff\\chapter 13\\exampleconvert.xlsx")
sheet = wb['Sheet1']
sheet.max_row
sheet.max_column
